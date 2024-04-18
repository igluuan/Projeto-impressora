import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
from tkinter import ttk

ctk.set_appearance_mode("System") 
appWidth, appHeight = 350, 210

class PrinterManager(ctk.CTk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry(f"{appWidth}x{appHeight}")
        self.title("Gerenciador de Impressoras")

        self.label_modelo = ctk.CTkLabel(self, text="Modelo:") 
        self.label_modelo.grid(row=0, column=0, sticky='ew')

        self.entry_modelo = ctk.CTkEntry(self, width=170, height=35, placeholder_text='Ex: 4172')
        self.entry_modelo.grid(row=0, column=1, sticky='ew')

        self.label_patrimonio = ctk.CTkLabel(self, text="Patrimônio:")
        self.label_patrimonio.grid(row=1, column=0,sticky='ew')

        self.entry_patrimonio = ctk.CTkEntry(self, width=170, height=35, placeholder_text='Ex: 12234')
        self.entry_patrimonio.grid(row=1, column=1,sticky='ew')

        self.label_marca = ctk.CTkLabel(self, text="Marca:")
        self.label_marca.grid(row=2, column=0,sticky='ew')

        # Alteração da entrada de marca para uma lista de seleção
        self.label_marca = tk.Label(self, text="Marca:")  # Use Label from tkinter
        self.label_marca.grid(row=2, column=0, sticky='ew')

        # Using tkinter's Combobox instead
        self.marca_options = ["SAMSUNG", "OKIDATA", "HP"]
        self.entry_marca = ttk.Combobox(self, values=self.marca_options, height=35)
        self.entry_marca.grid(row=2, column=1, sticky='ew')
        
        self.button_adicionar = ctk.CTkButton(self, text="Adicionar Impressora", command=self.adicionar_impressora, width=200, height=35)
        self.button_adicionar.grid(row=3, column=0, columnspan=2,pady=10,padx=20,sticky='ew')

        self.button_remover = ctk.CTkButton(self, text="Remover Impressora", command=self.remover_impressora, width=200, height=35)
        self.button_remover.grid(row=4, column=0, columnspan=2,padx=20, sticky='ew')

    def adicionar_impressora(self):
        modelo = self.entry_modelo.get()
        patrimonio = self.entry_patrimonio.get()
        marca = self.entry_marca.get()

        if not modelo or not patrimonio or not marca:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        try:
            workbook = load_workbook("impressoras.xlsx")
        except FileNotFoundError:
            workbook = Workbook()
            workbook.remove(workbook.active) # type: ignore

    
        worksheet = workbook['QUINTO_SETOR']
        worksheet.append([marca, modelo, patrimonio])
        workbook.save("impressoras.xlsx")
        print('ok')

        # Limpar campos após adicionar
        self.entry_modelo.delete(0, ctk.END)
        self.entry_patrimonio.delete(0, ctk.END)
        self.entry_marca.set("")  # Limpar a seleção da marca após adicionar

    def remover_impressora(self):
        modelo = self.entry_modelo.get()
        patrimonio = self.entry_patrimonio.get()

        if not patrimonio:
            messagebox.showerror("Erro", "Digite o patrimônio da impressora que deseja remover!")
            return
        if not modelo:
            messagebox.showerror("Erro", "Digite o modelo da impressora que deseja remover!")
            return
        try:
            workbook = load_workbook("impressoras.xlsx")
            worksheet = workbook[modelo]
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row[1] == patrimonio:
                    worksheet.delete_rows(row_idx)
                    workbook.save("impressoras.xlsx")
                    messagebox.showinfo("Sucesso", f"Impressora com modelo {modelo} e patrimônio {patrimonio} removida com sucesso!")
                    # Limpar campo após remover
                    self.entry_patrimonio.delete(0, ctk.END)
                    return
        except FileNotFoundError:
            messagebox.showerror("Erro", "Nenhuma impressora encontrada!")
            return

        messagebox.showerror("Erro", f"Impressora com patrimônio {patrimonio} não encontrada!")

    def run(self):
        self.mainloop()

if __name__ == "__main__":
    manager = PrinterManager()
    manager.run()
